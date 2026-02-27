"""客户导入导出服务"""
import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import pandas as pd
from openpyxl.utils import get_column_letter

from app.models.customer.info import CustomerInfo
from app.models.customer.contact import ContactInfo
from app.schemas.customer.customer import CustomerCreate, CustomerQueryParams
from app.services.customer.customer_service import CustomerService
from app.core.utils.helpers import string_to_tags


class ImportExportService:
    """客户导入导出服务类"""

    @staticmethod
    def _safe_str(value: Any) -> Optional[str]:
        if pd.isna(value):
            return None
        value_str = str(value).strip()
        return value_str or None

    @staticmethod
    def _safe_int(value: Any) -> Optional[int]:
        if pd.isna(value):
            return None
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _parse_customer_type(raw: Optional[str]) -> str:
        value = (raw or "").strip().lower()
        if value in {"individual", "个人"}:
            return "individual"
        return "enterprise"

    @staticmethod
    def _map_level(raw_level: Optional[str], raw_level_3: Optional[str]) -> str:
        # 优先使用原有 level 列（A/B/C/D）
        if raw_level:
            level = raw_level.strip().upper()
            if level in {"A", "B", "C", "D"}:
                return level

        # 兼容 3#客户分级（常见为 1/2/3/4）
        level_3 = (raw_level_3 or "").strip()
        mapping = {"1": "A", "2": "B", "3": "C", "4": "D"}
        return mapping.get(level_3, "C")

    @staticmethod
    async def export_customers(
        db: AsyncSession,
        customer_ids: List[int] = None,
        params: CustomerQueryParams = None,
    ) -> bytes:
        """
        导出客户数据为Excel

        Args:
            db: 数据库会话
            customer_ids: 指定导出的客户ID列表（优先级高）
            params: 查询参数（当未指定customer_ids时使用）

        Returns:
            Excel文件的二进制数据
        """
        # 获取客户数据
        if customer_ids:
            result = await db.execute(
                select(CustomerInfo).where(
                    CustomerInfo.id.in_(customer_ids),
                    CustomerInfo.deleted_at.is_(None)
                )
            )
            customers = result.scalars().all()
        elif params:
            customers, _ = await CustomerService.get_customers_paginated(db, params)
        else:
            # 默认导出所有客户
            result = await db.execute(
                select(CustomerInfo).where(CustomerInfo.deleted_at.is_(None))
            )
            customers = result.scalars().all()

        # 准备导出数据
        export_data = []
        for customer in customers:
            # 获取主要联系人
            contact_result = await db.execute(
                select(ContactInfo).where(
                    ContactInfo.customer_id == customer.id,
                    ContactInfo.is_primary.is_(True),
                    ContactInfo.deleted_at.is_(None)
                )
            )
            primary_contact = contact_result.scalar_one_or_none()

            row = {
                "客户编号": customer.customer_no,
                "1#客户名称（中）": customer.customer_name,
                "1#客户名称（EN）": customer.customer_name_en or "",
                "区域": customer.region or "",
                "3#客户类型": customer.customer_type_3 or "",
                "3#客户分级": customer.customer_level_3 or "",
                "5#成交客户": customer.deal_customer_5 if customer.deal_customer_5 is not None else "",
                "5#电气工程师人数": customer.electrical_engineer_count_5 if customer.electrical_engineer_count_5 is not None else "",
                "3#负责人": customer.owner_name_3 or customer.owner_name or "",
                "客户名称": customer.customer_name,
                "客户类型": "企业" if customer.customer_type == "enterprise" else "个人",
                "所属行业": customer.industry or "",
                "公司规模": customer.company_size or "",
                "法人代表": customer.legal_person or "",
                "注册资本": str(customer.registered_capital) if customer.registered_capital else "",
                "成立日期": customer.establish_date.strftime("%Y-%m-%d") if customer.establish_date else "",
                "省份": customer.province or "",
                "城市": customer.city or "",
                "区县": customer.district or "",
                "详细地址": customer.address or "",
                "官方网站": customer.website or "",
                "公司信息": customer.company_info or "",
                "产品信息": customer.product_info or "",
                "客户来源": customer.source or "",
                "客户级别": customer.level or "",
                "状态": customer.status or "",
                "负责人ID": customer.owner_id or "",
                "标签": customer.tags or "",
                "备注": customer.remarks or "",
                # 联系人信息
                "联系人姓名": primary_contact.name if primary_contact else "",
                "联系人职位": primary_contact.title if primary_contact else "",
                "联系人手机": primary_contact.mobile if primary_contact else "",
                "联系人邮箱": primary_contact.email if primary_contact else "",
                "创建时间": customer.created_at.strftime("%Y-%m-%d %H:%M:%S") if customer.created_at else "",
            }
            export_data.append(row)

        # 生成Excel文件
        df = pd.DataFrame(export_data)

        # 使用ExcelWriter生成xlsx文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='客户数据', index=False)

            # 调整列宽
            worksheet = writer.sheets['客户数据']
            for idx, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[get_column_letter(idx)].width = min(max_length + 2, 50)

        output.seek(0)
        return output.read()

    @staticmethod
    async def import_customers(
        db: AsyncSession,
        file_data: bytes,
        creator_id: int,
        owner_name: str = None,
    ) -> Dict[str, Any]:
        """
        从Excel导入客户数据

        Args:
            db: 数据库会话
            file_data: Excel文件的二进制数据
            creator_id: 创建人ID
            owner_name: 负责人姓名

        Returns:
            导入结果统计
        """
        # 读取Excel文件
        df = pd.read_excel(io.BytesIO(file_data), sheet_name=0)

        # 标准化列名
        column_mapping = {
            "1#客户名称（中）": "customer_name",
            "1#客户名称（EN）": "customer_name_en",
            "区域": "region",
            "3#客户类型": "customer_type_3",
            "3#客户分级": "customer_level_3",
            "5#成交客户": "deal_customer_5",
            "5#电气工程师人数": "electrical_engineer_count_5",
            "3#负责人": "owner_name_3",
            "客户名称": "customer_name",
            "客户类型": "customer_type",
            "所属行业": "industry",
            "公司规模": "company_size",
            "法人代表": "legal_person",
            "注册资本": "registered_capital",
            "成立日期": "establish_date",
            "省份": "province",
            "城市": "city",
            "区县": "district",
            "详细地址": "address",
            "官方网站": "website",
            "公司信息": "company_info",
            "产品信息": "product_info",
            "客户来源": "source",
            "客户级别": "level",
            "标签": "tags",
            "备注": "remarks",
        }
        df.rename(columns=column_mapping, inplace=True)

        # 导入统计
        success_count = 0
        error_count = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                # 跳过空行
                if pd.isna(row.get("customer_name")):
                    continue

                customer_type_value = ImportExportService._safe_str(row.get("customer_type"))
                customer_type_3_value = ImportExportService._safe_str(row.get("customer_type_3"))
                level_value = ImportExportService._safe_str(row.get("level"))
                level_3_value = ImportExportService._safe_str(row.get("customer_level_3"))
                row_owner_name_3 = ImportExportService._safe_str(row.get("owner_name_3"))

                # 准备客户数据
                customer_data = CustomerCreate(
                    customer_name=str(row["customer_name"]).strip(),
                    customer_name_en=ImportExportService._safe_str(row.get("customer_name_en")),
                    region=ImportExportService._safe_str(row.get("region")),
                    customer_type_3=customer_type_3_value,
                    customer_level_3=level_3_value,
                    deal_customer_5=ImportExportService._safe_int(row.get("deal_customer_5")),
                    electrical_engineer_count_5=ImportExportService._safe_int(row.get("electrical_engineer_count_5")),
                    owner_name_3=row_owner_name_3,
                    customer_type=ImportExportService._parse_customer_type(customer_type_value or customer_type_3_value),
                    industry=ImportExportService._safe_str(row.get("industry")),
                    company_size=ImportExportService._safe_str(row.get("company_size")),
                    legal_person=ImportExportService._safe_str(row.get("legal_person")),
                    registered_capital=float(row.get("registered_capital")) if pd.notna(row.get("registered_capital")) else None,
                    province=ImportExportService._safe_str(row.get("province")),
                    city=ImportExportService._safe_str(row.get("city")),
                    district=ImportExportService._safe_str(row.get("district")),
                    address=ImportExportService._safe_str(row.get("address")),
                    website=ImportExportService._safe_str(row.get("website")),
                    company_info=ImportExportService._safe_str(row.get("company_info")),
                    product_info=ImportExportService._safe_str(row.get("product_info")),
                    source=ImportExportService._safe_str(row.get("source")),
                    level=ImportExportService._map_level(level_value, level_3_value),
                    tags=string_to_tags(str(row.get("tags")).strip()) if pd.notna(row.get("tags")) else [],
                    remarks=ImportExportService._safe_str(row.get("remarks")),
                )

                # 创建客户
                await CustomerService.create_customer(
                    db,
                    customer_data,
                    creator_id=creator_id,
                    owner_name=row_owner_name_3 or owner_name,
                )
                success_count += 1

            except Exception as e:
                error_count += 1
                errors.append({
                    "row": idx + 2,  # Excel行号（从1开始，加表头）
                    "data": row.to_dict(),
                    "error": str(e),
                })

        return {
            "success_count": success_count,
            "error_count": error_count,
            "errors": errors,
        }

    @staticmethod
    def get_import_template() -> bytes:
        """
        生成客户导入模板

        Returns:
            Excel模板文件的二进制数据
        """
        # 定义模板列
        template_data = {
            "区域": ["华东苏皖", "华南"],
            "1#客户名称（中）": ["示例企业有限公司", "示例科技有限公司"],
            "1#客户名称（EN）": ["Example Industry Co., Ltd.", "Example Tech Co., Ltd."],
            "3#客户类型": ["B", "C"],
            "3#客户分级": ["1", "3"],
            "5#成交客户": [1, 0],
            "5#电气工程师人数": [80, 25],
            "3#负责人": ["张三", "李四"],
        }

        df = pd.DataFrame(template_data)

        # 生成Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='客户数据', index=False)

            # 调整列宽
            worksheet = writer.sheets['客户数据']
            for idx, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[get_column_letter(idx)].width = min(max_length + 2, 50)

        output.seek(0)
        return output.read()
